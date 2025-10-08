import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side

# Try to import pyxlsb for .xlsb files
try:
    import pyxlsb
    HAS_PYXLSB = True
except ImportError:
    HAS_PYXLSB = False
    print("Warning: pyxlsb not installed. Cannot read .xlsb files. Install with: pip install pyxlsb")

# Base directory where downloads are placed
FRANCE_FILES_DIR = Path(r"C:\Users\il00030293\OneDrive - Sysco Corporation\Documents\PGM\France files")

# Resolve latest Lumpsums file (pattern: "Lumpsums - vYYYY.MM.DD.xlsx")
def get_latest_lumpsums_file(base_dir: Path) -> Path:
    candidates = sorted(
        base_dir.glob("Lumpsums - v*.xlsb"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"No Lumpsums workbook found in {base_dir}")
    return candidates[0]

latest_input = get_latest_lumpsums_file(FRANCE_FILES_DIR)
input_file_name = str(latest_input)
output_file_name = str(latest_input.with_stem(latest_input.stem + "_output"))
sheet_name = 'Lumpsums 2025'

def deduplicate_and_preserve_format(input_file_name, output_file_name, sheet_name):
    # If input is .xlsb, convert to .xlsx first
    if input_file_name.endswith('.xlsb'):
        if not HAS_PYXLSB:
            print(f"Error: Cannot process .xlsb files without pyxlsb. Install with: pip install pyxlsb")
            return
        
        # Convert .xlsb to .xlsx
        temp_xlsx = input_file_name.replace('.xlsb', '_temp.xlsx')
        try:
            # Read with pyxlsb and save as xlsx
            df = pd.read_excel(input_file_name, engine='pyxlsb', sheet_name=sheet_name)
            df.to_excel(temp_xlsx, index=False, engine='openpyxl', sheet_name=sheet_name)
            input_file_name = temp_xlsx
            print(f"Converted .xlsb to temporary .xlsx: {temp_xlsx}")
        except Exception as e:
            print(f"Error converting .xlsb to .xlsx: {e}")
            return
    
    wb = load_workbook(input_file_name)
    ws = wb[sheet_name]
    
    # 1) Extract data into a DataFrame, using the 2nd row as headers
    data = ws.values
    first_row = next(data)   # skip the first row (maybe a title, etc.)
    headers = next(data)     # the second row becomes the DataFrame's header
    df = pd.DataFrame(data, columns=headers)
    
    # 2) Check the column to split
    col_to_split = 'Code Article / Code Sous-Gamme'
    if col_to_split not in df.columns:
        raise KeyError(f"Column '{col_to_split}' not found in the file.")
    
    # 3) Build a new set of rows after splitting ( ',' or '-' )
    result_rows = []
    for _, row in df.iterrows():
        cell_val = str(row[col_to_split]) if pd.notna(row[col_to_split]) else ''
        split_values = [v.strip() for v in re.split('[,-]', cell_val) if v.strip()]
        
        if len(split_values) > 1:
            for value in split_values:
                new_row = row.copy()
                new_row[col_to_split] = value
                result_rows.append(new_row)
        else:
            result_rows.append(row)
    
    # 4) Build a final DataFrame
    df_result = pd.DataFrame(result_rows)
    
    # 5) Clear the old content (without clearing existing styles) starting from row 3
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    
    # 6) Reinsert the new data into the sheet
    for r_idx, row_data in enumerate(dataframe_to_rows(df_result, index=False, header=True), start=2):
        for c_idx, value in enumerate(row_data, start=1):
            if pd.isna(value) or value == "None":
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 7) Fill columns A and H with the desired formulas (IFERROR to avoid #N/A)
    min_data_row = 3
    max_data_row = 2 + len(df_result)
    for row_idx in range(min_data_row, max_data_row + 1):
        # Column A => encapsulate VLOOKUP in IFERROR
        ws.cell(row=row_idx, column=1).value = (
            f"=IFERROR(VLOOKUP($D{row_idx},Réf!$A:$B,2,0),\"\")"
        )
        # Column H => same
        ws.cell(row=row_idx, column=8).value = (
            f"=IFERROR(VLOOKUP($G{row_idx},Réf!$D:$E,2,0),\"\")"
        )
    
    # 8) Apply formatting (font, alignment, formats, etc.)
    
    # 8.a) Calibri font, size 10 for the whole data range
    for row_idx in range(min_data_row, max_data_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Calibri', size=10)
    
    # 8.b) Center columns: B, E, F, G, I, Q
    centered_columns = ['B','E','F','G','I','Q']
    for col_letter in centered_columns:
        for row_idx in range(min_data_row, max_data_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.alignment = Alignment(horizontal='center')
    
    # 8.c) Columns J, R, T in currency format
    #      => 3-section format usage: positive; negative; zero => "-"
    currency_columns = ['J','R','T']
    custom_currency_format = '#,##0.00\\ "€";-#,##0.00\\ "€";"-   €"'  
    # -> If the value is 0, Excel will display "-"
    for col_letter in currency_columns:
        for row_idx in range(min_data_row, max_data_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = custom_currency_format
    
    # 8.d) Column S in percentage format
    for row_idx in range(min_data_row, max_data_row + 1):
        cell = ws[f"S{row_idx}"]
        cell.number_format = '0.00%'
    
    # 9) Add the requested vertical borders
    thin_side = Side(border_style="thin", color="000000")
    medium_side = Side(border_style="medium", color="000000")
    
    borders_to_apply = {
        # Thin borders
        'A': 'thin',  # between A and B
        'C': 'thin',  # between C and D
        'H': 'thin',  # between H and I
        'I': 'thin',  # between I and J
        'Q': 'thin',  # between Q and R
        'R': 'thin',  # between R and S
        'S': 'thin',  # between S and T
        
        # "medium" borders
        'J': 'medium', # between J and K
        'K': 'medium', # between K and Q
        'T': 'medium', # between T and U
    }
    
    for row_idx in range(min_data_row, max_data_row + 1):
        for col_letter, border_type in borders_to_apply.items():
            cell = ws[f"{col_letter}{row_idx}"]
            current_border = cell.border
            if border_type == 'thin':
                new_right_side = thin_side
            else:
                new_right_side = medium_side
            
            cell.border = Border(
                left=current_border.left,
                right=new_right_side,
                top=current_border.top,
                bottom=current_border.bottom
            )
    
    # 10) Final save
    wb.save(output_file_name)
    print(f"File saved successfully to: {output_file_name}")
    
    # 11) Clean up temporary file if it was created
    if input_file_name.endswith('_temp.xlsx'):
        try:
            Path(input_file_name).unlink()
            print(f"Cleaned up temporary file: {input_file_name}")
        except Exception as e:
            print(f"Warning: Could not delete temporary file {input_file_name}: {e}")

deduplicate_and_preserve_format(input_file_name, output_file_name, sheet_name)
